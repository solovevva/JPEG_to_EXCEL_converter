print('КОНВЕРТЕР: JPG -> XLSX')
print('by Victor Solovev')
print('e-mail: victor.solovev@gmail.com')
print()
print('ВАЖНО! Для корректной работы в папке с программой должен нахоиться файл input.jpg')
print()
print('Программа выпоняется')

#подключаю библиотеку для работы с изображением
from PIL import Image

#подключаю библиотеку для работы с MS Excel
from openpyxl import Workbook
from openpyxl.styles import PatternFill  # Подключаем стили для ячеек
from openpyxl.utils import get_column_letter # Функция для преобразования цифрового обозначение столбуа в буквенное

image = Image.open("input.jpg") #Открываем изображение.

#меняю размер изображения
image = image.resize((int(280 / (image.size[1] / image.size[0])), 280))

image = image.rotate(90, expand=True) #Поворот изображения на 90 градусов
image = image.transpose(Image.FLIP_TOP_BOTTOM) #Зеркальное отображение изображения относительно вертикальной оси
width = image.size[0] #Определяем ширину.
height = image.size[1] #Определяем высоту.
pix = image.load() #Выгружаем значения пикселей.

wb = Workbook() # Создали книгу
work_sheet = wb.active # Создали лист с названием и сделали его активным

#делаем ячейки квадратного размера путем изменения их ширины
for i in range(height):
    work_sheet.column_dimensions[get_column_letter(i+1)].width = 3

#перерисовываем пиксели картинки в заливку ячеек
log = width * height - 1
for i in range(width):
    for j in range(height):
        a = pix[i, j][0] * 6 // 255 << 5
        b = pix[i, j][1] * 6 // 255 << 5
        c = pix[i, j][2] * 6 // 255 << 5
        d = '%02x%02x%02x' % (a, b, c)
        work_sheet_a1 = work_sheet.cell(row=(i + 1), column = (j + 1))
        work_sheet_a1.fill = PatternFill(fill_type='solid', start_color=d,
                                 end_color=d)  # Данный код позволяет делать оформление цветом ячейки
        print('\r','Осталось обработать пикселей:', log, end='')
        log -= 1
print()
print('Создаю файл output.xlsx...')
wb.save("output.xlsx")
print('Файл output.xlsx создан. Выполнение программы завершено.')